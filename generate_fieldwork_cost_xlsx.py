"""Generate an Excel workbook to budget a single field trip.

This script creates a multi-sheet Excel template with transparent formulas
and named rates suitable for planning costs under typical Baden Wuerttemberg (BW)
framing as used in DFG-funded projects. Running the module as a script writes the
workbook to the current working directory.

Usage
-----
Run the script directly::

    python generate_fielwork_cost_xlsx.py

This produces ``/output/fieldtrip-cost-template.xlsx``. Open the file and
fill in **Inputs & Rates** first; all other sheets reference those values.

Sheets created
--------------
- **Inputs & Rates**: Centralized parameters (per-diem full/partial,
  private-car rate and trip-cap, rental per-km, default overnight,
  Hiwi hourly rate).
- **Staff & Participants**: One row per person. Roles supported:
  WiMi, Lab (VA), Hiwi (student assistant), Student (unpaid). Calculates
  per-diems (no meal deductions), overnights, hours-based wages for Hiwis,
  and participant subtotals.
- **Hours Log**: Date, task, first/last name, hours. Aggregated back to
  **Staff & Participants** via ``SUMIFS`` (names must match exactly).
- **Travel & Vehicles**: Tickets/day-rates, rental-car variable costs
  (km x per-km), private-car reimbursement (km x rate). The private-car
  trip-level cap is applied in **Summary**.
- **Materials & Other**: Consumables, equipment rentals, shipping,
  permits, and similar items.
- **Summary**: Category subtotals (per-diems, overnights, Hiwi wages,
  travel, materials) and a grand total. Applies the private-car cap.

Assumptions
-----------
The template uses configurable defaults aligned with common BW
practice (e.g., full/partial per-diems, private-car per-km with a
trip-level cap). Always replace defaults with the values required by
your institution and the specific grant documents.

Dependencies
------------
- Python 3.9+ (recommended)
- ``openpyxl`` for workbook creation

Side effects
------------
Writes ``/output/fieldtrip-cost-template.xlsx`` to the current directory
when executed as a script.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

def add_defined_name(wb, sheet_title: str, name: str, a1_ref: str) -> None:
    """
    Create/replace a workbook-level defined name pointing to a cell/range.
    Works across openpyxl variants where defined_names is list-like or dict-like.
    """
    target = f"'{sheet_title}'!{a1_ref}" if " " in sheet_title else f"{sheet_title}!{a1_ref}"

    delete = getattr(wb.defined_names, "delete", None)
    if callable(delete):
        try:
            delete(name)
        except Exception:
            pass
    else:
        try:
            if name in wb.defined_names:
                del wb.defined_names[name]
        except Exception:
            for dn in list(getattr(wb.defined_names, "__iter__", lambda: [])()):
                if getattr(dn, "name", None) == name:
                    try:
                        wb.defined_names.remove(dn)
                    except Exception:
                        pass

    dn = DefinedName(name=name, attr_text=target)
    if hasattr(wb.defined_names, "append"):
        wb.defined_names.append(dn)
    elif hasattr(wb.defined_names, "__setitem__"):
        wb.defined_names[name] = dn
    elif hasattr(wb.defined_names, "add"):
        wb.defined_names.add(dn)
    else:
        raise RuntimeError("Unsupported openpyxl defined_names container")


wb = Workbook()

# ------------------------------
# Styles & helpers
# ------------------------------
header_fill = PatternFill("solid", fgColor="F2F2F2")
bold = Font(bold=True)
thin = Side(style="thin", color="CCCCCC")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def make_header(row):
    for c in row:
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")
        c.border = border_all

def set_col_width(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

currency = NamedStyle(name="currency")
currency.number_format = '#,##0.00_);[Red](#,##0.00)'
percent = NamedStyle(name="percent")
percent.number_format = '0%'
try:
    wb.add_named_style(currency)
except ValueError:
    pass
try:
    wb.add_named_style(percent)
except ValueError:
    pass

# ------------------------------
# Sheet: Inputs & Rates
# ------------------------------
rates = wb.active
rates.title = "Inputs & Rates"

rates.append(["Item", "Value (EUR)", "Notes"])
make_header(rates[1])

rate_rows = [
    ("Per diem - full day (domestic)", 24, "Defaults to EUR 24 per full day"),
    ("Stadtmobil base rate", 150, "Set to average base rate in EUR for all Stadtmobil rentals"),
    ("Number of Stadtmobil cars", 1, "Define number of rented cars"),
    ("Total trip kilometers", 100, "Define the sum of all km per car"),
    ("Stadtmobil per-km cost (incl. fuel)", 0.35, "Planning value only; adjust to Stadtmobil rate for relevant cars"),
    ("Alternative - Stadtmobil lump sum", 0, "Enter a lump sum in EUR for all Stadtmobil rentals (e.g., for post-cost assessment)"),
    ("Default overnight cost per night", 95, "Planning cap or expected average incl. taxes; edit per trip"),
    ("Hiwi hourly rate (default)", 20.00, "Accounts for future wage raises"),
]
for r in rate_rows:
    rates.append(list(r))

# Style numeric cells
for r in range(2, 2+len(rate_rows)):
    label = str(rates.cell(row=r, column=1).value).lower()
    if any(k in label for k in ["per diem", "base rate", "overnight", "per-km", "hourly rate", "lump sum"]):
        rates.cell(row=r, column=2).style = "currency"

set_col_width(rates, [40, 18, 70])

# Named ranges for easy formulas (pass CELL A1 refs only!)
for nm, a1 in {
    "PER_DIEM": "$B$2",
    "STADTMOBIL_BASE": "$B$3",
    "STADTMOBIL_CAR_NUMBER": "$B$4",
    "TOTAL_KM": "$B$5",
    "STADTMOBIL_PER_KM": "$B$6",
    "STADTMOBIL_LUMPSUM": "$B$7",
    "OVERNIGHT_DEFAULT": "$B$8",
    "HIWI_RATE": "$B$9",
}.items():
    add_defined_name(wb, "Inputs & Rates", nm, a1)

# ------------------------------
# Sheet: Staff & Participants
# ------------------------------
staff = wb.create_sheet("Staff & Participants")
staff_headers = [
    "First name","Last name","Role (WiMi/VA/Hiwi/Unpaid graduating student)",
    "Trip start (date/time)","Trip end (date/time)",
    "Full-day count","Partial-day count (>8h or arr/dep)",
    "Per-diem total (EUR)",
    "Nights","Overnight cost per night (EUR)","Overnight total (EUR)",
    "Hours (from Hours Log)","Hourly rate (EUR)","Wages total (EUR)",
    "Participant subtotal (EUR)"
]
staff.append(staff_headers); make_header(staff[1])
set_col_width(staff, [16,16,30,20,20,16,22,18,10,22,18,22,16,12,18,20])

# Role dropdown
dv_role = DataValidation(
    type="list",
    formula1='"WiMi,Lab (VA),Hiwi (student assistant),Student (unpaid)"',
    allow_blank=True
)
staff.add_data_validation(dv_role)
dv_role.add("C2:C300")

for row in range(2, 301):
    # Per-diem total:
    # - WiMi & Hiwi: normal per-diem (full + partial, same rate as per your setup)
    # - Student (unpaid): NO per-diem -> 0
    staff.cell(row=row, column=8).value = (
        f"=IF(C{row}=\"Student (unpaid)\",0,IFERROR(F{row}*PER_DIEM + G{row}*PER_DIEM,0))"
    )
    staff.cell(row=row, column=8).style = "currency"

    # Overnight default and total
    staff.cell(row=row, column=10).value = "=OVERNIGHT_DEFAULT"
    staff.cell(row=row, column=10).style = "currency"
    staff.cell(row=row, column=11).value = f"=IFERROR(I{row}*J{row},0)"
    staff.cell(row=row, column=11).style = "currency"

    # Hours auto-summed from Hours Log by first & last name
    staff.cell(row=row, column=12).value = (
        f"=IFERROR(SUMIFS('Hours Log'!$F$2:$F$1000,'Hours Log'!$B$2:$B$1000,A{row},'Hours Log'!$C$2:$C$1000,B{row}),0)"
    )

    # Hiwi hourly rate; others 0
    staff.cell(row=row, column=13).value = f"=IF(C{row}=\"Hiwi (student assistant)\",HIWI_RATE,0)"
    staff.cell(row=row, column=13).style = "currency"


    # Wages total = Hours * Rate
    staff.cell(row=row, column=15).value = f"=IFERROR(L{row}*M{row},0)"
    staff.cell(row=row, column=15).style = "currency"

    # Subtotal = Per-diem + Overnight + Wages
    staff.cell(row=row, column=16).value = f"=H{row}+K{row}+O{row}"
    staff.cell(row=row, column=16).style = "currency"

# Totals row
tot_row = 302
staff.cell(row=tot_row, column=7, value="Totals:").font = bold
for c in (8,11,12,15,16):
    staff.cell(row=tot_row, column=c, value=f"=SUM({get_column_letter(c)}2:{get_column_letter(c)}301)")
    if c != 12:
        staff.cell(row=tot_row, column=c).style = "currency"

staff.freeze_panes = "A2"

# ------------------------------
# Sheet: Hours Log
# ------------------------------
hours = wb.create_sheet("Hours Log")
hours_headers = ["Date","Task/Activity","First name","Last name","Role (opt.)","Hours"]
hours.append(hours_headers); make_header(hours[1])
set_col_width(hours, [12,36,16,16,18,10])

dv_hours = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
hours.add_data_validation(dv_hours)
dv_hours.add("F2:F1000")
hours.freeze_panes = "A2"

# ------------------------------
# Sheet: Travel & Vehicles
# ------------------------------
travel = wb.create_sheet("Travel & Vehicles")
travel_headers = [
    "Date","Type (Train/Flight/Rental/Private/Taxi/PT)","Route / Purpose / Notes",
    "Ticket/Day rate (EUR)","Qty (days / tickets)","Line item (EUR)",
    "Rental km (estimate)","Rental per-km (EUR)","Rental variable (EUR)",
    "Private-car km","Private-car reimb. (EUR)","Travel subtotal (EUR)"
]
travel.append(travel_headers); make_header(travel[1])
set_col_width(travel, [12,18,36,16,14,16,18,16,16,16,18,18])

dv_type = DataValidation(type="list", formula1='"Train,Flight,Rental,Private,Taxi,Public Transport,Other"', allow_blank=True)
travel.add_data_validation(dv_type)
dv_type.add("B2:B500")

for row in range(2, 501):
    # Keep line items for tickets/day-rates
    travel.cell(row=row, column=6).value = f"=IFERROR(D{row}*E{row},0)"
    travel.cell(row=row, column=6).style = "currency"

    # We compute Stadtmobil centrally in Summary via the named rates; avoid double counting here:
    travel.cell(row=row, column=8).value = ""           # Rental per-km left blank
    travel.cell(row=row, column=9).value = "=0"         # Rental variable suppressed
    travel.cell(row=row, column=11).value = "=0"        # Private car suppressed

    travel.cell(row=row, column=12).value = f"=F{row}+I{row}+K{row}"
    travel.cell(row=row, column=12).style = "currency"

t_tot = 502
travel.cell(row=t_tot, column=3, value="Totals:").font = bold
for c in (6,9,11,12):
    travel.cell(row=t_tot, column=c, value=f"=SUM({get_column_letter(c)}2:{get_column_letter(c)}501)")
    travel.cell(row=t_tot, column=c).style = "currency"
travel.freeze_panes = "A2"

# ------------------------------
# Sheet: Material Expenses
# ------------------------------
other = wb.create_sheet("Material Expenses")
other_headers = [
    "Date","Item / Description","Category (consumables/equipment/shipping/permits/other)",
    "Units","Unit cost (EUR)","Line total (EUR)","Notes"
]
other.append(other_headers); make_header(other[1])
set_col_width(other, [12,30,38,10,16,16,30])

for row in range(2, 401):
    other.cell(row=row, column=6).value = f"=IFERROR(D{row}*E{row},0)"
    other.cell(row=row, column=6).style = "currency"

o_tot = 402
other.cell(row=o_tot, column=3, value="Totals:").font = bold
other.cell(row=o_tot, column=6, value=f"=SUM(F2:F401)").style = "currency"
other.freeze_panes = "A2"

# ------------------------------
# Sheet: Summary
# ------------------------------
summary = wb.create_sheet("Summary")
set_col_width(summary, [40,22,28])

summary.merge_cells("A1:C1")
summary["A1"] = "Field Trip Cost Summary"
summary["A1"].font = Font(bold=True, size=14)
summary["A1"].alignment = Alignment(horizontal="center")

summary.append(["","",""])  # spacer
summary.append(["Category","Subtotal (EUR)","Notes"]); make_header(summary[3])

# Staff/participants subtotals
summary.append(["Per-diems (total)", "=IFERROR('Staff & Participants'!H302,0)", "WiMi, VA & Hiwi only; unpaid students excluded."])
summary.append(["Overnights", "=IFERROR('Staff & Participants'!K302,0)", "Nights x cost/night."])
summary.append(["Hiwi wages", "=IFERROR('Staff & Participants'!O302,0)", "Hours x rate if used)."])

# Travel tickets (non-Stadtmobil)
summary.append(["Tickets / day-rates (travel)", "=IFERROR('Travel & Vehicles'!F502,0)", "Trains, flights, taxis, PT, etc."])

# Stadtmobil (central calc; uses your new variables)
summary.append([
    "Stadtmobil (cars, base + km) or lump sum",
    "=IF(STADTMOBIL_LUMPSUM>0, STADTMOBIL_LUMPSUM, STADTMOBIL_CAR_NUMBER*STADTMOBIL_BASE + TOTAL_KM*STADTMOBIL_PER_KM)",
    "If a lump sum is provided (>0), it overrides the calculated cost."
])

# Materials & other
summary.append(["Materials & other", "=IFERROR('Material Expenses'!F402,0)", "Consumables, rentals, permits, shipping."])

# Grand total
summary.append(["Grand total (EUR)", "=SUM(B4:B9)", "Includes Stadtmobil and all other categories."])

# Currency styling
for r in range(4, 11):
    summary.cell(row=r, column=2).style = "currency"
summary.cell(row=10, column=1).font = Font(bold=True)
summary.cell(row=10, column=2).font = Font(bold=True)
summary.cell(row=10, column=2).style = "currency"

# Notes
summary.append(["","",""])
summary.append(["Notes","","Set rates in 'Inputs & Rates'. Roles: WiMi, VA staff & Hiwis may receive per-diem; 'Student (unpaid)' receives overnights only."])

summary.freeze_panes = "A4"

# Save workbook (ensure folder exists)
out_path = Path("output/fieldtrip-cost-template.xlsx")
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(out_path)
print(out_path)
